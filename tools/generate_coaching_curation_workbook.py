"""Generate the coaching/front-office curation workbook.

Output: data/scheme/curation/coaching_curation.xlsx — four tabs:
  • Head Coach
  • Offensive Coordinator
  • Defensive Coordinator
  • General Manager

Pre-fills team names + 1-2 quality-bar example rows per tab. Brett
fills the rest using the schema documented in
docs/scheme_curation_guide.md (and inline header comments).

Re-run safely; overwrites the file.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
OUT = REPO / "data" / "scheme" / "curation" / "coaching_curation.xlsx"

TEAMS = [
    "ARI", "ATL", "BAL", "BUF", "CAR", "CHI", "CIN", "CLE",
    "DAL", "DEN", "DET", "GB",  "HOU", "IND", "JAX", "KC",
    "LA",  "LAC", "LV",  "MIA", "MIN", "NE",  "NO",  "NYG",
    "NYJ", "PHI", "PIT", "SEA", "SF",  "TB",  "TEN", "WAS",
]

HEADER_FILL = PatternFill("solid", fgColor="0a3d62")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FILL = PatternFill("solid", fgColor="0076B6")
TITLE_FONT = Font(bold=True, color="FFFFFF", size=14)
EXAMPLE_FILL = PatternFill("solid", fgColor="FFF8DC")
THIN = Side(style="thin", color="B0B7BC")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Reasonable as-of-2024-staff seeds; Brett will verify currency.
HC_SEEDS = {
    "ARI": "Jonathan Gannon",   "ATL": "Raheem Morris",
    "BAL": "John Harbaugh",     "BUF": "Sean McDermott",
    "CAR": "Dave Canales",      "CHI": "Ben Johnson",  # 2025 hire
    "CIN": "Zac Taylor",        "CLE": "Kevin Stefanski",
    "DAL": "Brian Schottenheimer",  # 2025 hire
    "DEN": "Sean Payton",       "DET": "Dan Campbell",
    "GB":  "Matt LaFleur",      "HOU": "DeMeco Ryans",
    "IND": "Shane Steichen",    "JAX": "Liam Coen",  # 2025 hire
    "KC":  "Andy Reid",         "LA":  "Sean McVay",
    "LAC": "Jim Harbaugh",      "LV":  "Pete Carroll",  # 2025 hire
    "MIA": "Mike McDaniel",     "MIN": "Kevin O'Connell",
    "NE":  "Mike Vrabel",       "NO":  "Kellen Moore",  # 2025 hire
    "NYG": "Brian Daboll",      "NYJ": "Aaron Glenn",   # 2025 hire
    "PHI": "Nick Sirianni",     "PIT": "Mike Tomlin",
    "SEA": "Mike Macdonald",    "SF":  "Kyle Shanahan",
    "TB":  "Todd Bowles",       "TEN": "Brian Callahan",
    "WAS": "Dan Quinn",
}


def _make_header_row(ws, row_idx, headers, fill=HEADER_FILL, font=HEADER_FONT):
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER


def _set_widths(ws, widths):
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def _add_title(ws, title, n_cols):
    ws.merge_cells(start_row=1, start_column=1,
                     end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center",
                                  vertical="center")


def _fill_team_rows(ws, start_row, hc_seeds=None,
                       example_team=None, example_values=None):
    """Pre-populate 32 team rows starting at row `start_row`."""
    for i, t in enumerate(TEAMS):
        r = start_row + i
        ws.cell(row=r, column=1, value=t)
        if hc_seeds and t in hc_seeds:
            ws.cell(row=r, column=2, value=hc_seeds[t])
        # Example row gets full pre-fill + highlight
        if t == example_team and example_values:
            for col_idx, val in enumerate(example_values, start=2):
                ws.cell(row=r, column=col_idx, value=val)
            for col_idx in range(1, len(example_values) + 2):
                ws.cell(row=r, column=col_idx).fill = EXAMPLE_FILL


def _build_hc_tab(wb):
    ws = wb.create_sheet("Head Coach")
    headers = [
        "Team", "Name", "Years in role",
        "System / identity",
        "Coaching tree / school",
        "Mentor (primary)",
        "Career path",
        "Leadership archetype",
        "4th-down aggression",
        "Game-management rep",
        "Acts as OC?",
        "Acts as DC?",
        "Signature win / moment",
        "One-liner identity",
        "Notes",
    ]
    n_cols = len(headers)
    _add_title(ws, "🏈  HEAD COACH — current staff", n_cols)
    _make_header_row(ws, 2, headers)
    _set_widths(ws, [
        8, 22, 8, 30, 28, 22, 38, 28, 18, 38,
        12, 12, 38, 50, 30,
    ])
    ws.freeze_panes = "C3"

    example = (
        "Dan Campbell",  # Name
        "5 (2021-)",     # Years
        "Defensive-minded culture-builder, offensive empowerment",  # System
        "Belichick tree (DC) → Saints culture (Payton)",  # Tree
        "Sean Payton",   # Mentor
        "MIA TE → DAL → NO TE coach → DET HC",  # Path
        "Players' coach, locker-room first",  # Leadership
        "Aggressive (top-3 in 4th-down go rate)",  # 4th-down
        "Sometimes too aggressive in -EV spots; "
        "owns mistakes publicly",  # Game mgmt
        "No",   # Acts as OC?
        "No",   # Acts as DC?
        "2024 NFC North title; 15-2 regular season",  # Sig moment
        "Lions HC who turned culture around with toughness "
        "and offensive empowerment of Ben Johnson.",  # One-liner
        "",  # Notes
    )
    _fill_team_rows(ws, 3, HC_SEEDS,
                       example_team="DET", example_values=example)
    return ws


def _build_oc_tab(wb):
    ws = wb.create_sheet("Offensive Coordinator")
    headers = [
        "Team", "Name", "Years in role",
        "Architect status",
        "Passing system",
        "Running system",
        "Personnel preference",
        "Pace identity",
        "QB type designed for",
        "Red-zone identity",
        "Coaching tree",
        "Mentor (primary)",
        "Signature concepts",
        "OL coach (architect of run game)",
        "One-liner identity",
        "Notes",
    ]
    n_cols = len(headers)
    _add_title(ws, "📔  OFFENSIVE COORDINATOR — current staff "
                   "(if HC is architect, use HC name + status="
                   "\"HC = architect\")", n_cols)
    _make_header_row(ws, 2, headers)
    _set_widths(ws, [
        8, 22, 8, 22,
        24, 26, 22, 18, 18, 22, 22, 22, 50,
        22, 50, 30,
    ])
    ws.freeze_panes = "C3"

    example = (
        "John Morton",   # Name (2025 OC, replaced Ben Johnson)
        "1 (2025-)",
        "Co-coordinator (Campbell trusts him; BJ template)",
        "WCO (Ben Johnson template inheritance)",
        "Outside zone, with power-counter complement",
        "11-heavy (90%+)",
        "Methodical with situational tempo",
        "Pocket passer (Goff)",
        "Creative (BJ legacy of trickery)",
        "Shanahan tree (via Ben Johnson)",
        "Ben Johnson (philosophical inheritance)",
        "Pre-snap motion · Condensed sets · "
        "Outside zone · Play-action · TE-heavy red zone",
        "Hank Fraley (zone scheme)",
        "Continuity hire — preserve the Ben Johnson template "
        "while Campbell leads from the HC chair.",
        "VERIFY: 2025 OC may differ; was BJ through 2024.",
    )
    _fill_team_rows(ws, 3, None,
                       example_team="DET", example_values=example)
    return ws


def _build_dc_tab(wb):
    ws = wb.create_sheet("Defensive Coordinator")
    headers = [
        "Team", "Name", "Years in role",
        "Architect status",
        "Coverage scheme",
        "Front structure",
        "Pressure tendencies",
        "Pass-rush plan",
        "Run defense identity",
        "Critical-situation tendency",
        "Coaching tree",
        "Mentor (primary)",
        "Signature concepts",
        "One-liner identity",
        "Notes",
    ]
    n_cols = len(headers)
    _add_title(ws, "🛡️  DEFENSIVE COORDINATOR — current staff "
                   "(if HC is architect, use HC name + status="
                   "\"HC = architect\")", n_cols)
    _make_header_row(ws, 2, headers)
    _set_widths(ws, [
        8, 22, 8, 22, 26, 18, 26, 26, 24, 32, 22, 22, 50, 50, 30,
    ])
    ws.freeze_panes = "C3"

    example = (
        "Kelvin Sheppard",  # 2025 DC after Glenn left for NYJ
        "1 (2025-)",
        "Dedicated",
        "Cover-3 base / pattern-match (AG inheritance)",
        "Multiple (4-3 base, sub-package heavy)",
        "Disguised pressure, 4-man rush primary",
        "Stunts/twists; Hutchinson speed off the edge",
        "Gap-control, attack-style",
        "Cover-2 in red zone; rare 3rd-and-long blitz",
        "Glenn tree (Saints/Falcons/Lions lineage)",
        "Aaron Glenn (philosophical inheritance)",
        "Disguised single-high · Mug looks · "
        "Hutchinson/AA edge speed · Hybrid LB roles",
        "Continuity hire after Glenn → NYJ HC. "
        "Preserves the Glenn defensive identity.",
        "VERIFY: 2025 DC pending; was AG through 2024.",
    )
    _fill_team_rows(ws, 3, None,
                       example_team="DET", example_values=example)
    return ws


def _build_gm_tab(wb):
    ws = wb.create_sheet("General Manager")
    headers = [
        "Team", "Name", "Years in role",
        "Roster-build philosophy",
        "Cap allocation",
        "Position-group priority",
        "Draft tendencies",
        "Risk tolerance",
        "Mentor / school",
        "Career path",
        "Notable hits / misses",
        "One-liner identity",
        "Notes",
    ]
    n_cols = len(headers)
    _add_title(ws, "💼  GENERAL MANAGER — current staff", n_cols)
    _make_header_row(ws, 2, headers)
    _set_widths(ws, [
        8, 22, 8, 26, 24, 26, 24, 22, 26, 38, 50, 50, 30,
    ])
    ws.freeze_panes = "C3"

    example = (
        "Brad Holmes",
        "5 (2021-)",
        "Draft-and-develop with selective FA splash",
        "Top-heavy stars + drafted depth",
        "OL-first / DL second",
        "BPA with positional value tilt to trenches",
        "Production over measurables; "
        "willing on character risks",
        "Howie Roseman (philosophical) / Rams scout dept",
        "STL/LA Rams scout (1995-2020) → DET GM",
        "Hits: Hutchinson, Sewell, Gibbs, LaPorta, ARSB. "
        "Misses: Okwara extension, Levi Onwuzurike",
        "Built the Lions' transformation through scouting "
        "depth and trench investment.",
        "",
    )
    _fill_team_rows(ws, 3, None,
                       example_team="DET", example_values=example)
    return ws


def main() -> None:
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _build_hc_tab(wb)
    _build_oc_tab(wb)
    _build_dc_tab(wb)
    _build_gm_tab(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"✓ wrote {OUT.relative_to(REPO)}")
    print(f"  4 tabs: Head Coach / Offensive Coordinator / "
          f"Defensive Coordinator / General Manager")
    print(f"  32 team rows per tab, alphabetical")
    print(f"  DET pre-filled as quality-bar example on each tab")
    print(f"  HC names seeded with 2024-2025 best-known staff "
          f"(verify currency)")


if __name__ == "__main__":
    main()
